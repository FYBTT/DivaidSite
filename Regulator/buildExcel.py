# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import pickle
import os
import sys
reload(sys)
sys.setdefaultencoding('utf-8')
goods = dict()

wb = load_workbook('./ExcelList/cutting nozzle.xlsx')
sheet_names = wb.get_sheet_names()
for inc in xrange(len(sheet_names)):
    name = sheet_names[inc]    
    ws = wb.get_sheet_by_name(name)
    lines = []
    speStart = 0
    lineNum = 0
    for row in ws.iter_rows(min_row = ws.min_row, min_col = 5, max_col = 10, max_row = ws.max_row):
        line = []
        for cell in row:
            if cell.value:
#                if cell.value == 'Specifications:':
#                    speStart = lineNum
                line.append(str(cell.value))
        lines.append(line)
        lineNum = lineNum + 1
    linesNum = len(lines)
    spe = dict()
    feaVector = []
    lines[2] = lines[2][1:]
    lines[3] = lines[3][1:]
    lenSpe = len(lines[2])
    for i in xrange(lenSpe):
        re = []
        for line in lines[3:]:
            re.append(line[i])
        spe[lines[2][i]] = re
#    for line in lines[speStart + 1:]:
#        spe[line[0]] = line[1]
#    for line in lines[3: speStart]:
#        if len(line) == 0:
#            continue
#        fea = line[0][2:]
#        feaVector.append(fea)
    goodType = 'cutting nozzle'
    countryID = lines[0][0].split('-')[-1][0:2]
    goodDict = dict()
    goodDict['order_no'] = lines[0][0]
    goodDict['typeInShort'] = goodType
    goodDict['country'] = countryID
    goodDict['typeDetail'] = lines[1][0]
    goodDict['feaList'] = feaVector
    goodDict['specification'] = spe
    goods[lines[0][0]] = goodDict        
    print lines

