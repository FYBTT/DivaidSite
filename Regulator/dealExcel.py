# -*- coding: utf-8 -*-
from openpyxl import load_workbook
specification = open("result.csv", 'wb')
relation = open('relation.csv', 'wb')
relation.write('goods_id, feature_id\n')
specification.write('order_no,gas,max_inlet_pressure,max_outlet_pressure,inlet_connection,outlet_connection,Type\n')
feature = open("feature.csv", 'wb')
wb = load_workbook('regulator.xlsx')
sheet_names = wb.get_sheet_names()
feaDict = dict()
feaCount = 1
for inc in xrange(len(sheet_names)):
    name = sheet_names[inc]
    print name
    specification.write(str(inc + 1))
    specification.write(',')
    ws = wb.get_sheet_by_name(name)
    line = []
    for row in ws.iter_rows(min_row = ws.max_row, min_col = 5, max_col = 10, max_row = ws.max_row):
        for cell in row:
            line.append(str(cell.value))
    fea = []
    for row in ws.iter_rows(min_row = 4, min_col = 5, max_col = 5, max_row = ws.max_row - 3):
        for cell in row:
            if cell.value in feaDict:
                relation.write(str(inc+1)+','+ str(feaDict[cell.value]))
                relation.write('\n')
            else:
                feaDict[cell.value] = feaCount
                relation.write(str(inc+1)+ ','+ str(feaDict[cell.value]))
                relation.write('\n')
                feature.write(str(feaCount))
                feature.write(',')
                feaCount = feaCount + 1
                feature.write(cell.value[2:-1])
                feature.write('\n')
#    line.append('_'.join(fea))
    line.append(ws['E2'].value)
    specification.write(','.join(line))
    specification.write('\n')
specification.close()
feature.close()